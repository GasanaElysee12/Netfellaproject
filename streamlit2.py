import re
import json
import pandas as pd
import numpy as np
import spacy
import streamlit as st
from PyPDF2 import PdfReader
import openpyxl
import tempfile
import os
import difflib
from io import StringIO
import base64
import en_core_web_sm

# Environment fixes
os.environ['TF_ENABLE_ONEDNN_OPTS'] = '0'
os.environ['STREAMLIT_SERVER_FILE_WATCHER_TYPE'] = 'none'

# ---------------------------
# 1. DATA PREPARATION MODULE
# ---------------------------
class TaxonomyBuilder:
    def __init__(self, excel_path):
        self.excel_path = excel_path
        self.unspsc_map = {}
        self.attribute_vocab = {}
        
    def build_taxonomy(self):
        """Build hierarchical taxonomy from Excel structure"""
        df = pd.read_excel(self.excel_path,skiprows=12)
        
        # Create UNSPSC code mapping
        for _, row in df.iterrows():
            if not pd.isna(row['Commodity']):
                code = str(int(row['Commodity']))
                self.unspsc_map[code] = {
                    'segment': str(int(row['Segment'])) if not pd.isna(row['Segment']) else '',
                    'segment_title': row['Segment Title'],
                    'family': str(int(row['Family'])) if not pd.isna(row['Family']) else '',
                    'family_title': row['Family Title'],
                    'class': str(int(row['Class'])) if not pd.isna(row['Class']) else '',
                    'class_title': row['Class Title'],
                    'commodity_title': row['Commodity Title'],
                    'definition': row['Commodity Definition'],
                    'synonyms': self._parse_synonyms(row.get('Synonym', ''))
                }
        
        # Build attribute vocabulary (simplified example)
        self.attribute_vocab = {
            "15121503": {  # Fuel tanks
                "capacity": ["500L", "1000L", "2000L"],
                "material": ["Steel", "Plastic"],
                "type": ["Diesel", "Petrol"],
                "feature": ["With Pump", "Insulated", "Portable"]
            },
            "24111601": {  # Gloves
                "type": ["Latex", "Nitrile", "Vinyl"],
                "size": ["Small", "Medium", "Large", "XL"],
                "use": ["Surgical", "Industrial", "Examination"],
                "sterility": ["Sterile", "Non-Sterile"]
            },
            "24101601": {  # Tractors
                "type": ["Farm", "Agricultural"],
                "power": ["50HP", "75HP", "100HP"],
                "use": ["Ploughing", "Harvesting", "Hauling"],
                "fuel": ["Diesel", "Electric", "Gasoline"]
            }
        }
        
        return self.unspsc_map, self.attribute_vocab
    
    def _parse_synonyms(self, synonym_str):
        """Parse synonyms from Excel cell"""
        if pd.isna(synonym_str):
            return []
        return [s.strip() for s in str(synonym_str).split(';') if s.strip()]

# ---------------------------
# 2. INPUT PROCESSING MODULE
# ---------------------------
class InputProcessor:
    def __init__(self):
        # Load spaCy model
        self.nlp = en_core_web_sm.load()
    
    def process_input(self, input_data):
        """Process various input formats to extract text"""
        if isinstance(input_data, str):
            # Single text input
            return self._clean_text(input_data)
        
        elif isinstance(input_data, list):
            # Batch processing - list of product descriptions
            return [self._clean_text(item) for item in input_data]
        
        elif hasattr(input_data, 'name') and input_data.name.endswith('.pdf'):
            # PDF file
            return self._extract_pdf_text(input_data)
        
        elif hasattr(input_data, 'name') and input_data.name.endswith(('.xlsx', '.xls')):
            # Excel file
            return self._extract_excel_text(input_data)
        
        elif isinstance(input_data, dict):
            # JSON input
            return self._extract_json_text(input_data)
        
        else:
            raise ValueError("Unsupported input format")
    
    def _clean_text(self, text):
        """Clean and normalize text"""
        text = re.sub(r'[^\w\s.,;:]', '', text)  # Keep basic punctuation
        text = re.sub(r'\s+', ' ', text).strip()
        return text.lower()
    
    def _extract_pdf_text(self, pdf_file):
        """Extract text from PDF"""
        text = ""
        reader = PdfReader(pdf_file)
        for page in reader.pages:
            text += page.extract_text() + "\n"
        return self._clean_text(text)
    
    def _extract_excel_text(self, excel_file):
        """Extract text from Excel"""
        text = ""
        with tempfile.NamedTemporaryFile(delete=False) as tmp:
            tmp.write(excel_file.getvalue())
            tmp_path = tmp.name
        
        wb = openpyxl.load_workbook(tmp_path)
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            for row in sheet.iter_rows(values_only=True):
                for cell in row:
                    if cell and isinstance(cell, str):
                        text += cell + " "
        
        os.unlink(tmp_path)
        return self._clean_text(text)
    
    def _extract_json_text(self, json_data):
        """Extract text from JSON structure"""
        def extract_values(obj):
            if isinstance(obj, dict):
                for v in obj.values():
                    yield from extract_values(v)
            elif isinstance(obj, list):
                for item in obj:
                    yield from extract_values(item)
            elif isinstance(obj, str):
                yield obj
        
        text = " ".join(extract_values(json_data))
        return self._clean_text(text)

# ---------------------------
# 3. UNSPSC CLASSIFICATION MODULE (Rule-based)
# ---------------------------
class UNSPSCClassifier:
    def __init__(self, unspsc_map):
        self.unspsc_map = unspsc_map
        self.keyword_mapping = self._build_keyword_mapping()
        
    def _build_keyword_mapping(self):
        """Create mapping from keywords to UNSPSC codes"""
        mapping = {}
        for code, info in self.unspsc_map.items():
            keywords = [info['commodity_title'].lower()]
            keywords.extend(syn.lower() for syn in info.get('synonyms', []))
            for kw in keywords:
                mapping[kw] = code
        
        # Add additional mappings
        mapping.update({
            "fuel": "15121503",
            "tank": "15121503",
            "glove": "24111601",
            "tractor": "24101601",
            "agriculture": "24101601",
            "diesel": "15121503",
            "surgical": "24111601",
            "farm": "24101601"
        })
        return mapping
    
    def predict(self, description):
        """Predict UNSPSC code using keyword matching"""
        # First try exact keyword matching
        for keyword, code in self.keyword_mapping.items():
            if keyword in description:
                return code
        
        # Then try fuzzy matching if no exact match
        best_match = None
        best_score = 0
        
        for keyword, code in self.keyword_mapping.items():
            score = difflib.SequenceMatcher(None, keyword, description).ratio()
            if score > best_score and score > 0.6:
                best_score = score
                best_match = code
        
        return best_match if best_match else "00000000"  # Default code

# ---------------------------
# 4. ATTRIBUTE EXTRACTION MODULE
# ---------------------------
class AttributeExtractor:
    def __init__(self, attribute_vocab):
        self.attribute_vocab = attribute_vocab
        
    def extract_attributes(self, description, unspsc_code):
        """Extract attributes from product description"""
        attributes = {}
        
        # Check if we have attribute vocabulary for this commodity
        if unspsc_code not in self.attribute_vocab:
            return attributes
        
        # Look for attribute values in text
        for attr, possible_values in self.attribute_vocab[unspsc_code].items():
            for value in possible_values:
                if value.lower() in description:
                    attributes[attr] = value
                    break
            else:
                # Try to find patterns for known attributes
                if attr == "capacity":
                    match = re.search(r'(\d+)\s*(L|liters?|litres?)', description, re.IGNORECASE)
                    if match:
                        attributes["capacity"] = f"{match.group(1)}L"
                
                elif attr == "power":
                    match = re.search(r'(\d+)\s*(HP|horsepower)', description, re.IGNORECASE)
                    if match:
                        attributes["power"] = f"{match.group(1)}HP"
                
                elif attr == "size":
                    # Look for size indicators
                    for size in ["small", "medium", "large", "xl"]:
                        if size in description:
                            attributes["size"] = size.capitalize()
                            break
                
                elif attr == "sterility":
                    if "sterile" in description:
                        attributes["sterility"] = "Sterile"
                    elif "non-sterile" in description or "non sterile" in description:
                        attributes["sterility"] = "Non-Sterile"
        
        return attributes

# ---------------------------
# 5. ATTRIBUTE STANDARDIZATION MODULE
# ---------------------------
class AttributeStandardizer:
    def __init__(self, attribute_vocab):
        self.attribute_vocab = attribute_vocab
    
    def standardize(self, attributes, unspsc_code):
        """Standardize extracted attribute values"""
        if unspsc_code not in self.attribute_vocab:
            return attributes
        
        standardized = {}
        for attr, value in attributes.items():
            if attr in self.attribute_vocab[unspsc_code]:
                # Find closest match in possible values
                possible_values = self.attribute_vocab[unspsc_code][attr]
                best_match = difflib.get_close_matches(
                    value, 
                    possible_values, 
                    n=1, 
                    cutoff=0.7
                )
                
                if best_match:
                    standardized[attr] = best_match[0]
                else:
                    standardized[attr] = value
            else:
                standardized[attr] = value
        
        return standardized

# ---------------------------
# 6. HUMAN-IN-THE-LOOP INTERFACE
# ---------------------------
class HILTInterface:
    def __init__(self, unspsc_map, attribute_vocab):
        self.unspsc_map = unspsc_map
        self.attribute_vocab = attribute_vocab
    
    def display_interface(self, original_desc, unspsc_code, extracted_attrs, standardized_attrs):
        """Display validation interface"""
        st.subheader("AI Classification Results")
        st.write(f"**Original Description:** `{original_desc}`")
        
        # Display UNSPSC classification
        st.subheader("Commodity Classification")
        if unspsc_code in self.unspsc_map:
            commodity_info = self.unspsc_map[unspsc_code]
            st.write(f"**Code:** {unspsc_code}")
            st.write(f"**Title:** {commodity_info['commodity_title']}")
            st.write(f"**Definition:** {commodity_info['definition']}")
        else:
            st.warning(f"Unknown UNSPSC code: {unspsc_code}")
        
        # Display attribute extraction
        st.subheader("Extracted Specifications")
        corrected_attrs = {}
        
        if unspsc_code in self.attribute_vocab:
            for attr in self.attribute_vocab[unspsc_code].keys():
                col1, col2 = st.columns([1, 3])
                with col1:
                    st.write(f"**{attr.capitalize()}:**")
                
                with col2:
                    current_value = standardized_attrs.get(attr, "")
                    options = self.attribute_vocab[unspsc_code][attr]
                    
                    if options:
                        # Create select box with options
                        selected = st.selectbox(
                            f"Select {attr}",
                            options=options,
                            index=options.index(current_value) if current_value in options else 0,
                            key=f"attr_{attr}_{id(original_desc)}"
                        )
                        corrected_attrs[attr] = selected
                    else:
                        # Free-form text input
                        corrected_attrs[attr] = st.text_input(
                            f"{attr.capitalize()}", 
                            value=current_value,
                            key=f"attr_{attr}_{id(original_desc)}"
                        )
        else:
            st.warning(f"No attribute taxonomy available for commodity {unspsc_code}")
        
        # Feedback mechanism
        st.subheader("Feedback")
        feedback = st.text_area("Provide feedback on the classification (optional)", 
                              key=f"feedback_{id(original_desc)}")
        
        if st.button("Submit Correction", key=f"submit_{id(original_desc)}"):
            return corrected_attrs, feedback
        
        return None, None

# ---------------------------
# 7. MAIN PROCESSING PIPELINE
# ---------------------------
class CommodityClassificationSystem:
    def __init__(self, excel_path):
        # Build taxonomies
        self.taxonomy_builder = TaxonomyBuilder(excel_path)
        self.unspsc_map, self.attribute_vocab = self.taxonomy_builder.build_taxonomy()
        
        # Initialize modules
        self.input_processor = InputProcessor()
        self.unspsc_classifier = UNSPSCClassifier(self.unspsc_map)
        self.attribute_extractor = AttributeExtractor(self.attribute_vocab)
        self.attribute_standardizer = AttributeStandardizer(self.attribute_vocab)
        self.hilt_interface = HILTInterface(self.unspsc_map, self.attribute_vocab)
        
        # Feedback database
        self.feedback_db = []
    
    def process_input(self, input_data):
        """Full processing pipeline for single or multiple inputs"""
        # Extract and clean text
        clean_data = self.input_processor.process_input(input_data)
        
        # Handle single input
        if isinstance(clean_data, str):
            return self._process_single(clean_data)
        # Handle batch processing
        elif isinstance(clean_data, list):
            results = []
            for desc in clean_data:
                results.append(self._process_single(desc))
            return results
        else:
            return self._process_single(clean_data)
    
    def _process_single(self, description):
        """Process a single product description"""
        # Classify UNSPSC code
        unspsc_code = self.unspsc_classifier.predict(description)
        
        # Extract attributes
        extracted_attrs = self.attribute_extractor.extract_attributes(description, unspsc_code)
        
        # Standardize attributes
        standardized_attrs = self.attribute_standardizer.standardize(
            extracted_attrs, unspsc_code
        )
        
        return {
            "original": description,
            "clean_description": description,
            "unspsc_code": unspsc_code,
            "extracted_attributes": extracted_attrs,
            "standardized_attributes": standardized_attrs,
            "commodity_title": self.unspsc_map.get(unspsc_code, {}).get("commodity_title", "Unknown")
        }
    
    def handle_feedback(self, result, corrected_attrs, feedback):
        """Store user feedback for model improvement"""
        self.feedback_db.append({
            "original_input": result["original"],
            "ai_classification": result["unspsc_code"],
            "ai_attributes": result["standardized_attributes"],
            "corrected_attributes": corrected_attrs,
            "user_feedback": feedback
        })
    
    def _get_table_download_link(self, df):
        """Generate a download link for a DataFrame"""
        csv = df.to_csv(index=False)
        b64 = base64.b64encode(csv.encode()).decode()
        return f'<a href="data:file/csv;base64,{b64}" download="classification_results.csv">Download CSV</a>'

    def web_interface(self):
        """Streamlit-based web interface"""
        st.title("AI-Powered Commodity Classification System")
        st.markdown("""
        This system automatically classifies products using UNSPSC codes and extracts 
        standardized specifications from product descriptions.
        """)
        
        # Input method selection
        input_method = st.radio("Select input method:", 
                               ["Single Product", "Batch Processing", "PDF", "Excel", "JSON"])
        
        input_data = None
        
        if input_method == "Single Product":
            input_data = st.text_area("Enter product description:", height=150, key="text_input")
            
        elif input_method == "Batch Processing":
            batch_input = st.text_area("Enter one product per line:", height=300, key="batch_input")
            if batch_input:
                input_data = [line.strip() for line in batch_input.split('\n') if line.strip()]
                
        elif input_method == "PDF":
            pdf_file = st.file_uploader("Upload PDF file", type=["pdf"], key="pdf_uploader")
            if pdf_file:
                input_data = pdf_file
                
        elif input_method == "Excel":
            excel_file = st.file_uploader("Upload Excel file", type=["xlsx", "xls"], key="excel_uploader")
            if excel_file:
                input_data = excel_file
                
        elif input_method == "JSON":
            json_str = st.text_area("Enter JSON data:", height=150, key="json_input")
            if json_str:
                try:
                    input_data = json.loads(json_str)
                except json.JSONDecodeError:
                    st.error("Invalid JSON format")
        
        if st.button("Process Input", key="process_btn") and input_data:
            with st.spinner("Processing..."):
                try:
                    results = self.process_input(input_data)
                    
                    # Handle batch results differently
                    if isinstance(results, list):
                        st.success(f"Processed {len(results)} products!")
                        
                        # Create results table
                        results_data = []
                        for r in results:
                            row = {
                                "Description": r["original"],
                                "UNSPSC Code": r["unspsc_code"],
                                "Commodity": r["commodity_title"]
                            }
                            # Add attributes
                            for attr, value in r["standardized_attributes"].items():
                                row[attr.capitalize()] = value
                            results_data.append(row)
                            
                        results_df = pd.DataFrame(results_data)
                        
                        # Display results
                        st.subheader("Batch Classification Results")
                        st.dataframe(results_df)
                        
                        # Download button
                        st.markdown(self._get_table_download_link(results_df), unsafe_allow_html=True)
                    else:
                        # Single result display
                        st.success("Classification Complete!")
                        st.subheader("Results Summary")
                        
                        # Display UNSPSC classification
                        if results["unspsc_code"] in self.unspsc_map:
                            commodity_info = self.unspsc_map[results["unspsc_code"]]
                            st.markdown(f"**UNSPSC Code:** `{results['unspsc_code']}`")
                            st.markdown(f"**Commodity Title:** {commodity_info['commodity_title']}")
                            st.markdown(f"**Definition:** {commodity_info['definition']}")
                        else:
                            st.warning(f"Unknown UNSPSC code: {results['unspsc_code']}")
                        
                        # Display attributes
                        st.subheader("Standardized Attributes")
                        if results["standardized_attributes"]:
                            for attr, value in results["standardized_attributes"].items():
                                st.markdown(f"- **{attr.capitalize()}:** {value}")
                        else:
                            st.info("No attributes extracted")
                        
                        # Show HILT interface
                        st.divider()
                        st.subheader("Validate & Correct Results")
                        correction, feedback = self.hilt_interface.display_interface(
                            str(input_data)[:500] + "..." if len(str(input_data)) > 500 else str(input_data),
                            results["unspsc_code"],
                            results["extracted_attributes"],
                            results["standardized_attributes"]
                        )
                        
                        if correction:
                            self.handle_feedback(results, correction, feedback)
                            st.success("Thank you for your feedback! Correction submitted.")
                        
                except Exception as e:
                    st.error(f"Error processing input: {str(e)}")

# ---------------------------
# 8. APPLICATION ENTRY POINT
# ---------------------------
if __name__ == "__main__":
    # Configuration
    EXCEL_PATH = "./Data/unspsc-english-v260801.1.xlsx"  # Path to your Excel file
    
    # Initialize system
    system = CommodityClassificationSystem(EXCEL_PATH)
    
    # Start web interface
    system.web_interface()